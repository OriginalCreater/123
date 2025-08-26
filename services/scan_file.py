import os
# import time
from openpyxl import load_workbook, utils
from tkinter import filedialog as fd
from modules.search_pusk import one_affel_check, one_affel_email, one_affel_site, one_affel_name
# def path_file():
#     return fd.askopenfilename()


def scan_file(path):
    """Получение пути фаила с информацией 'ИНН'"""
    path = str(path)
    """Создание екземпляра класса для работы с екселем"""
    wb = load_workbook(filename=path)
    """Поиск страници по названию 'Основная'"""
    try:
        book = wb["Основная"]
    except:
        """rename активного листа в 'Основная'"""
        book = wb.active
        book.title = "Основная"
    """Создание списка с столбцами екселя"""
    cpi = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V",
           "W", "X", "Y", "Z"]
    """Создание списков для необходимых данных"""
    stolb_name = []
    string_range = []


    def scan_stolb_for_inn():
        """Поиск столбца с ИНН"""
        for number_stolb, name_stolb in enumerate(cpi):

            stolb = str(book[f'{name_stolb}1'].value)
            if stolb == "ИНН" or stolb == "Инн" or stolb == "инн":
                stolb_name.append(name_stolb)
                for string_inn in range(1, 1000):
                    """Поиск длинны строк с инн"""
                    max_inn = str(book[f'{name_stolb}{string_inn}'].value)
                    max_inn2 = str(book[f'{name_stolb}{string_inn+1}'].value)
                    max_inn3 = str(book[f'{name_stolb}{string_inn+2}'].value)
                    if max_inn == "None" and max_inn2 == "None" and max_inn3 == "None":
                        string_range.append(string_inn)
                        break
                """Создание столца для результатов проверки"""
                name_rezault_stolb = str(cpi[number_stolb + 1])
                name_reason = str(cpi[number_stolb + 2])
                wb.save(path)

                stolb_name.append(name_rezault_stolb)
                stolb_name.append(name_reason)
                rezault = str(book[f'{name_rezault_stolb}1'].value)
                print(f"Жма {rezault}")
                if rezault == "Результат (ПУСК)":
                    """Если найден столбез с резултатами поиск его длинны"""
                    for string_rezault in range(1, 1000):
                        """Поиск длинны колличества результатов для оценки минимального значения"""
                        max_rezault = str(book[f'{name_rezault_stolb}{string_rezault}'].value)
                        if max_rezault == "None":
                            """Запись длинны столбца с результатами"""
                            string_range.append(string_rezault)
                            break
                elif rezault == "None":
                    """Если столбец пуст rename его в 'Результат (ПУСК)'"""
                    book[f'{name_rezault_stolb}1'] = f"Результат (ПУСК)"
                    string_rezault = 2
                    string_range.append(string_rezault)
                else:
                    """Если чтото иное, создание нового столбца для результатов"""
                    book.insert_cols(number_stolb + 2, 2)
                    wb.save(path)
                    book[f'{name_rezault_stolb}1'] = f"Результат (ПУСК)"
                    book[f'{name_reason}1'] = f"Причина (ПУСК)"
                    string_rezault = 2
                    string_range.append(string_rezault)


    """Начало сканирования фаила на длинну и положение"""
    scan_stolb_for_inn()
    """Сохранение фаила"""
    wb.save(path)
    """Чтение максимальной длинны информации с длинной ИНН"""
    max_range = int(string_range[0])
    """Чтение макисмальной длинны информации с результатами"""
    min_range = int(string_range[1])
    """Сохранение положения столбца с ИНН"""
    inn_stolb = str(stolb_name[0])
    """Сохранение положения столбца с результатами"""
    rezault_stolb = str(stolb_name[1])
    reason_stolb = str(stolb_name[2])
    # print(f"Колличество ИНН: {max_range - 2}")
    # print(f"Колличество проверенных: {min_range - 2}")
    # print(f"ИНН нахоится в столбце: {inn_stolb}")
    # print(f"Результат находится на столбце: {rezault_stolb}")
    inn_list = []
    for i in range(min_range, max_range):
        inn = str(book[f'{inn_stolb}{i}'].value)
        inn = inn.split('.')
        inn = inn[0]
        print(inn)
        inn_list.append(inn)
    return inn_list, rezault_stolb, reason_stolb

def scan_file_affel(path, token):
    """Получение пути фаила с информацией 'ИНН'"""
    path = str(path)
    """Создание екземпляра класса для работы с екселем"""
    wb = load_workbook(filename=path)
    """Поиск страници по названию 'Основная'"""
    try:
        book = wb["Основная"]
    except:
        """rename активного листа в 'Основная'"""
        book = wb.active
        book.title = "Основная"
    """Создание списка с столбцами екселя"""
    cpi = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V",
           "W", "X", "Y", "Z"]
    """Создание списков для необходимых данных"""
    stolb_name = []
    string_range = []
    stolb_num = []
    phone_pos = []
    def scan_stolb_for_inn():
        """Поиск столбца с ИНН"""
        for number_stolb, name_stolb in enumerate(cpi):

            stolb = str(book[f'{name_stolb}1'].value)
            if stolb == "ИНН" or stolb == "Инн" or stolb == "инн":
                stolb_name.append(name_stolb)
                for string_inn in range(1, 1000):
                    """Поиск длинны строк с инн"""
                    max_inn = str(book[f'{name_stolb}{string_inn}'].value)
                    max_inn2 = str(book[f'{name_stolb}{string_inn+1}'].value)
                    max_inn3 = str(book[f'{name_stolb}{string_inn+2}'].value)
                    if max_inn == "None" and max_inn2 == "None" and max_inn3 == "None":
                        string_range.append(string_inn)
                        wb.save(path)
                        break
                """Создание столца для результатов проверки"""
                name_info_stolb = str(cpi[number_stolb+1])
                wb.save(path)
                stolb_num.append(name_info_stolb)
                rezault = str(book[f'{name_info_stolb}1'].value)
                print(rezault)
                if rezault == "Данные по аффилировке":
                    pass
                elif rezault == "None":
                    """Если столбец пуст rename его в 'Результат проверки'"""
                    book[f'{name_info_stolb}1'] = f"Данные по аффилировке"
                    string_rezault = 2
                    string_range.append(string_rezault)
                else:
                    """Если чтото иное, создание нового столбца для результатов"""
                    book.insert_cols(number_stolb + 2, 1)
                    wb.save(path)
                    book[f'{name_info_stolb}1'] = f"Данные по аффилировке"
                    string_rezault = 2
                    string_range.append(string_rezault)


    """Начало сканирования фаила на длинну и положение"""
    scan_stolb_for_inn()
    """Сохранение фаила"""
    wb.save(path)
    for number_stolb, name_stolb in enumerate(cpi):

        stolb = str(book[f'{name_stolb}1'].value)
        if stolb == "Телефоны":
            print(f"Номер на столбце - {name_stolb}")
            phone_pos.append(name_stolb)
    """Чтение максимальной длинны информации с длинной ИНН"""
    max_range = int(string_range[0])
    """Сохранение положения столбца с ИНН"""
    inn_stolb = str(stolb_name[0])
    """Сохранение столбца с номером """
    phone_stolb = str(phone_pos[0])
    """ХЗ"""
    num_stolb = str(stolb_num[0])
    inn_list = []
    for i in reversed(range(2, max_range)):
        # inn = str(book[f'{inn_stolb}{i}'].value)
        book.insert_rows(i)
        book[f'{inn_stolb}{i}'] = f"----------"
        wb.save(path)
    max_range_true = []
    for i in range(1, 1000):
        """Поиск длинны строк с инн"""
        max_inn = str(book[f'{inn_stolb}{i}'].value)
        max_inn2 = str(book[f'{inn_stolb}{i + 1}'].value)
        max_inn3 = str(book[f'{inn_stolb}{i + 2}'].value)
        max_inn4 = str(book[f'{inn_stolb}{i + 3}'].value)
        max_inn5 = str(book[f'{inn_stolb}{i + 4}'].value)
        if max_inn == "None" and max_inn2 == "None" and max_inn3 == "None" and max_inn4 == "None" and max_inn5 == "None":
            max_range_true.append(i)
            break
    max = max_range_true[0]
    print(max_range_true[0])
    list_info = []

    for i in range(1, max):
        inn = str(book[f'{inn_stolb}{i}'].value)
        if inn == f"----------":
            pass
        elif inn == "ИНН" or inn == "Инн" or inn == "инн":
            pass
        else:

            list_info.append(inn)
    def index_search(inn):
        for row_index in range(1, book.max_row + 1):  # Начинаем с 1 (первая строка)
            for col_index in range(1, book.max_column + 1):
                cell_value = book.cell(row=row_index, column=col_index).value
                if cell_value == int(inn) or cell_value == str(inn):  # Сравниваем с искомым значением
                    return row_index  # Возвращаем номер строки
    for i in list_info:
        # print(list_info)
        row_index_1 = index_search(i)
        print(i)
        row_index_1 = str(row_index_1)
        inn = row_index_1.split('.')
        row_index_1 = int(inn[0])

        # print(row_index_1)
        inn = str(book[f'{inn_stolb}{row_index_1}'].value)
        print(f"индекс = {row_index_1}")
        """phone_stolb = Буква столбца с инн, необходимо собрать номера по индексу и разложить на нужный формат 
                                                                   после чего засунуть в функцию с проверкой номеров"""
        list_phone = []
        phone = str(book[f'{phone_stolb}{row_index_1}'].value)

        # list_phone.append(phone)
        # print(list_phone)
        inn_list, num_list = one_affel_check(inn, token, row_index_1,phone)
        print(inn_list, num_list)
        inn_list_two, email_list = one_affel_email(inn, token, row_index_1)
        inn_list_tree, site_list = one_affel_site(inn, token, row_index_1)
        inn_list_four, name_list = one_affel_name(inn, token, row_index_1)
        # print(inn_list_two, email_list)
        if inn_list == False:
            book[f'{num_stolb}{row_index_1}'] = f"Проверить глазами"
            wb.save(path)
            continue
        else:
            pos = 0
            for s, a in enumerate(inn_list):
                if a == []:
                    pass
                else:
                    for i in range(row_index_1, row_index_1 + len(a)):
                        book.insert_rows(i + 1)
                        wb.save(path)
                        book[f'{inn_stolb}{i + 1}'] = f"{a[pos]}"
                        book[f'{num_stolb}{i + 1}'] = f"{num_list[s]}"
                        pos += 1
                        wb.save(path)
                    else:
                        pos = 0
        pos = 0
        print(inn_list_two, "\n", email_list)
        for d, a in enumerate(inn_list_two):
            if a == []:
                pass
            else:
                for i in range(row_index_1, row_index_1 + len(a)):
                    book.insert_rows(i + 1)
                    wb.save(path)

                    book[f'{inn_stolb}{i + 1}'] = f"{a[pos]}"
                    try:
                        book[f'{num_stolb}{i + 1}'] = f"{email_list[d]}"
                    except IndexError:
                        book[f'{num_stolb}{i + 1}'] = f"{email_list[-1]}"
                    wb.save(path)
                    pos += 1
                    wb.save(path)
                else:
                    pos = 0
        pos = 0
        for d, a in enumerate(inn_list_tree):
            if a == []:
                pass
            else:
                for i in range(row_index_1, row_index_1 + len(a)):
                    book.insert_rows(i + 1)
                    wb.save(path)
                    book[f'{inn_stolb}{i + 1}'] = f"{a[pos]}"
                    book[f'{num_stolb}{i + 1}'] = f"{site_list[d]}"
                    wb.save(path)
                    pos += 1
                    wb.save(path)
                else:
                    pos = 0
        for s, a in enumerate(inn_list_four):
            if a == []:
                pass
            else:
                for i in range(row_index_1, row_index_1 + len(a)):
                    book.insert_rows(i + 1)
                    wb.save(path)
                    book[f'{inn_stolb}{i + 1}'] = f"{a[pos]}"
                    book[f'{num_stolb}{i + 1}'] = f"{name_list[s]}"
                    pos += 1
                    wb.save(path)
                else:
                    pos = 0
    return inn_list

def duble_opti(path):
    """Получение пути фаила с информацией 'ИНН'"""
    path = str(path)
    """Создание екземпляра класса для работы с екселем"""
    wb = load_workbook(filename=path)
    """Поиск страници по названию 'Основная'"""
    try:
        book = wb["Основная"]
    except:
        """rename активного листа в 'Основная'"""
        book = wb.active
        book.title = "Основная"
    """Создание списка с столбцами екселя"""
    inn_data = []  # [(значение, строка, адрес_ячейки), ...]

    # Поиск столбца с ИНН
    inn_column = None
    for col in range(1, book.max_column + 1):
        cell_value = str(book.cell(row=1, column=col).value).lower()
        if cell_value == "ИНН" or cell_value == "инн" or cell_value == "Инн":
            inn_column = col
            break

    if inn_column:
        # Поиск всех ИНН в найденном столбце
        for row in range(2, book.max_row + 1):
            cell = book.cell(row=row, column=inn_column)
            cell_value = str(cell.value)

            if cell_value and cell_value != "None" and cell_value != "----------":
                cell_address = f"{chr(64 + inn_column)}{row}"  # Преобразуем номер в букву
                inn_data.append((cell_value, row, cell_address))

    # Разделяем данные по отдельным спискам
    inn = [item[0] for item in inn_data]
    poz = [item[1] for item in inn_data]
    str_poz = [item[2] for item in inn_data]

    print(f"Найдено ИНН: {len(inn)}")
    for i, (value, row, address) in enumerate(inn_data, 1):
        print(f"{i}. ИНН: {value}, Строка: {row}, Адрес: {address}")
        # print(inn, poz, str_poz)

    lst = inn
    # Исходный список
    print(f"Исходный список: {lst}")

    # Сбор статистики
    stats = {}
    for i, item in enumerate(lst):
        if item not in stats:
            stats[item] = {'count': 0, 'indexes': []}
        stats[item]['count'] += 1
        stats[item]['indexes'].append(i)

    # Вывод детальной информации
    print("\nДетальная статистика:")
    for item, data in stats.items():
        status = "ДУБЛИКАТ" if data['count'] > 1 else "УНИКАЛЬНЫЙ"
        print(f"'{item}': {data['count']} вхождений {status} - позиции {data['indexes']}")

    # Сбор индексов для удаления (все кроме первого вхождения)
    indexes_to_remove = []
    for item, data in stats.items():
        if data['count'] > 1:
            indexes_to_remove.extend(data['indexes'][1:])

    # Сортировка по возрастанию
    indexes_to_remove.sort()
    print(f"\nВсего дубликатов для удаления: {len(indexes_to_remove)}")
    print(f"Индексы для удаления: {indexes_to_remove}")
    poz_del = []
    for i in indexes_to_remove:
        poz_del.append(poz[i])
    if len(poz_del) == 0 or len(indexes_to_remove) == 0:
        pass
    else:
        poz_del.pop(-1)
        indexes_to_remove.pop(-1)
        print(f"Позиции для удаления: {poz_del}")
        # Удаление с конца
        print(f"\nПроцесс удаления:")
        for i, index in enumerate(reversed(indexes_to_remove), 1):
            value = lst[index]
            delit_pos = poz[index]
            print(f"{i:2d}. Удаляем '{value}' с позиции {delit_pos}")
            book.delete_rows(idx=delit_pos)
            del lst[index]
            wb.save(path)

        print(f"\nФинальный результат: {lst}")

# print(scan_file("C:\\Users\\evgeniy\\Desktop\\alve\\uploads\\1213425 (4) (2) (2).xlsx"))